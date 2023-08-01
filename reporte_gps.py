import pandas as pd
from openpyxl import Workbook
import openpyxl
import pyodbc

connection_string = "Driver={SQL Server};" \
                    "Server=192.168.1.00;" \
                    "Database=BD_TEST;" \
                    "UID=USER;" \
                    "PWD=PASS"


def gps_report(patent = '0', start_date = '20100101', end_date = '20300101'):
    saved_path = r'C:\Users\andres\reportes_gps\GPS_{}.xlsx'.format(patent)
    
    query = """
         SELECT CODIGO AS PATENTE,
                FECHAHORA AS FECHA,
                LAT AS LATITUD,
                LON AS LONGITUD,
                'https://maps.google.com/maps?q=' + CAST(LAT AS VARCHAR(50)) + ',' + CAST(LON AS VARCHAR(50)) AS 'UBICACIÓN',
             CASE
                WHEN ESTADO = 'A' THEN 'ANTENA VALIDA'
                ELSE 'INVALIDA'
             END AS 'ESTADO ANTENA',
             BATERIA AS 'BATERIA GPS',
             NUM_EVENTO AS EVENTO,
             CASE
                WHEN NUM_EVENTO = '40' THEN 'Bajo nivel batería'
                WHEN NUM_EVENTO = '50' THEN 'Desconexion alimentacion principal'
                WHEN NUM_EVENTO = '251' THEN 'Desconexion alimentacion principal'
                WHEN NUM_EVENTO = '28' THEN 'Motor apagado'
                WHEN NUM_EVENTO = '29' THEN 'Motor encendido'
                WHEN NUM_EVENTO = '45' THEN 'Informes con Ignición'
                WHEN NUM_EVENTO = '47' THEN 'Posición online / posición ignición off'
                WHEN NUM_EVENTO = '46' THEN 'Envío de datos de LOG'
                WHEN NUM_EVENTO = '31' THEN 'Logeo Ibutton'
                ELSE '-'
             END AS DESCRIPCION,
             CASE
                WHEN STD_IGN = 0 THEN 'APAGADO'
                ELSE 'ENCENDIDO'
             END AS IGNICION,
             VEL AS VELOCIDAD
         FROM MOV_POSICION WITH(NOLOCK)
         WHERE CODIGO IN ('{}') FECHAHORA BETWEEN '{}' AND '{}'
         ORDER BY 2 ASC""".format(patent, start_date, end_date)
    print(query)
    
    conn = pyodbc.connect(connection_string)

    try:
        dataframe = pd.read_sql_query(query, conn)
        dataframe.to_excel(saved_path, index=False)
        workbook = openpyxl.load_workbook(saved_path)
        
        sheet = workbook.active
        headers = dataframe.columns
        
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.number_format = '@'
        
        for j, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=2+j).value = header

        for i, row in enumerate(dataframe.values, start=2):
            for j, value in enumerate(row, start=2):
                sheet.cell(row=i, column=1+j).value = value
        
        
        print("EXCEL GUARDADO")
        
    except Exception as e:
        print("Error al ejecutar la consulta:", e)
        
    finally: 
        conn.close()   

#reporte GPS
gps_report('RTPX-39','20230730 00:00', '20230730 23:59')







